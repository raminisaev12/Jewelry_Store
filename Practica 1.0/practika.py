import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
import pymysql
from tkinter import messagebox
import random
from tkcalendar import DateEntry
from docx import Document
from tkinter import filedialog
from datetime import date
import sys
import os
import pymysql.cursors
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)



root = tk.Tk()
root.title("Алмазный путь")
root.geometry("600x400")
root.state('zoomed')
root.columnconfigure(0, weight=1)
root.columnconfigure(1, weight=1)
root.rowconfigure(2, weight=1)

l = Image.open(resource_path("logo2.png"))
l= ImageTk.PhotoImage(l)

root.iconphoto(True, l)

bg_color = "#FDF5E6"
text_color = "#2D6A4F"
root.configure(bg=bg_color)
root.rowconfigure(1, weight=1)

img = Image.open(resource_path("logo.png"))
img = img.resize((40, 40))
logo_img = ImageTk.PhotoImage(img)

# Логотип
frame = tk.Frame(root,bg=bg_color)
frame.grid(column=0, row=0, sticky="ew")

label_logo = tk.Label(frame, image=logo_img, bg=bg_color)
label_logo.grid(row=0, column=0, padx=(20, 5), pady=20, sticky="w")

# Заголовок
label_zagol = tk.Label(frame, text="Алмазный путь",
                       fg=text_color,
                       font=("Segoe UI", 22, "bold"),
                       bg=bg_color)
label_zagol.grid(row=0, column=0, padx=65, pady=20, sticky="w")

border_frame = tk.Frame(frame, bg=bg_color, highlightbackground="black",highlightcolor="black", highlightthickness=1)
border_frame.grid(row=1, column=0, padx=20, pady=2, sticky="w")

dobavl = tk.Label(border_frame, text="Добавление изделия", fg="#4A7C59",
                  font=("Segoe UI", 12, "italic"), bg=bg_color)
dobavl.grid(row=0, column=0, padx=10, pady=(5, 0), sticky="w")

##добавление
for i in range(6):
    border_frame.columnconfigure(i * 2, weight=0)  # Для Label
    border_frame.columnconfigure(i * 2 + 1, weight=1)  # Для Entry

####
def create_field(label_text, row_idx, col_idx):
    lbl = tk.Label(border_frame, text=label_text, bg=bg_color, fg=text_color)
    lbl.grid(row=row_idx, column=col_idx * 2, padx=(10, 5), pady=10, sticky="ew")

    ent = tk.Entry(border_frame, bg="white", relief="flat", highlightthickness=1, highlightbackground="#A0A0A0")
    ent.grid(row=row_idx, column=col_idx * 2 + 1, padx=(0, 10), pady=10, sticky="ew")
    return ent

def create_combobox(label_text, row_idx, col_idx, values_list):
    lbl = tk.Label(border_frame, text=label_text, bg=bg_color, fg=text_color)
    lbl.grid(row=row_idx, column=col_idx * 2, padx=(10, 5), pady=10, sticky="ew")

    combobox = ttk.Combobox(border_frame, values=values_list, state="readonly")
    combobox.grid(row=row_idx, column=col_idx * 2 + 1, padx=(0, 10), pady=10, sticky="ew")
    return combobox

def on_validate_text(new_value):
    """Разрешены буквы, пробел, дефис"""
    return all(c.isalpha() or c in (' ', '-') for c in new_value) if new_value else True

def on_validate_int(new_value):
    """Только цифры (целое неотрицательное)"""
    return new_value.isdigit() or new_value == ""

def on_validate_float(new_value):
    """Число с плавающей точкой (точка или запятая)"""
    if new_value == "":
        return True
    try:
        float(new_value.replace(',', '.'))
        return True
    except ValueError:
        return False

# Регистрируем функции в root
v_text = root.register(on_validate_text)
v_int = root.register(on_validate_int)
v_float = root.register(on_validate_float)


# Создаем поля по порядку
Entry_name = create_field("Название изделия", 1, 0)
Entry_name.config(validate="key", validatecommand=(v_text, '%P'))   #буквы/пробел/дефис

Entry_type = create_combobox("Тип", 1, 1, ["Кольцо", "Серьги", "Браслет", "Цепочка"])
Entry_category = create_combobox("Категория", 1, 2, ["Свадебные", "Женские", "Мужские", "Детские"])
Entry_metal = create_combobox("Металл", 1, 3, ["Золото", "Красное золото", "Белое золото", "Желтое золото", "Серебро"])
Entry_gemstone = create_combobox("Камень", 1, 4, ["Нет", "Бриллиант", "Сапфир", "Рубин", "Изумруд", "Фианит"])

Entry_purity = create_combobox("Проба", 2, 0, ["375", "500", "585", "750", "925", "958", "999"])

Entry_weight = create_field("Вес (г)", 2, 1)
Entry_weight.config(validate="key", validatecommand=(v_float, '%P')) # число с точкой/запятой

Entry_price = create_field("Цена (₽)", 2, 2)
Entry_price.config(validate="key", validatecommand=(v_float, '%P'))  #число


Entry_quantity = create_field("Количество", 2, 4)
Entry_quantity.config(validate="key", validatecommand=(v_int, '%P'))

lbl_size = tk.Label(border_frame, text="Размер", bg=bg_color, fg=text_color)
lbl_size.grid(row=2, column=3 * 2, padx=(10, 5), pady=10, sticky="ew")

Entry_size = ttk.Combobox(border_frame, values=["15", "15.5", "16", "16.5", "17", "17.5", "18"], state="readonly")
Entry_size.grid(row=2, column=3 * 2 + 1, padx=(0, 10), pady=10, sticky="ew")
# Функция для обновления списка размеров в зависимости от типа
def update_size_options(event):
    selected_type = Entry_type.get()

    if selected_type == "Серьги":
        lbl_size.config(text="Размер")
        Entry_size['values'] = ["—"]
        Entry_size.set("—")

    elif selected_type == "Кольцо":
        lbl_size.config(text="Размер")
        Entry_size['values'] = ["15", "15.5", "16", "16.5", "17", "17.5", "18", "18.5", "19"]
        Entry_size.set("17")

    elif selected_type == "Браслет":
        lbl_size.config(text="Обхват (см)")
        Entry_size['values'] = ["16", "17", "18", "19", "20", "21"]
        Entry_size.set("18")

    elif selected_type == "Цепочка":
        lbl_size.config(text="Длина (см)")
        Entry_size['values'] = ["40", "45", "50", "55", "60", "65"]
        Entry_size.set("45")

    else:
        lbl_size.config(text="Размер")
        Entry_size['values'] = ["—"]
        Entry_size.set("—")

Entry_type.bind("<<ComboboxSelected>>", update_size_options)
#скелет
cols = ('id', 'name', 'type', 'category', 'metal', 'gemstone', 'purity', 'weight', 'size', 'quantity', 'price')

tree = ttk.Treeview(root, columns=cols, show='headings',height=12)

tree.heading('id', text='ID',anchor="center")
tree.heading('name', text='Название изделия',anchor="center")
tree.heading('category', text='Категория',anchor="center")
tree.heading('type', text='Тип',anchor="center")
tree.heading('metal', text='Металл',anchor="center")
tree.heading('gemstone', text='Камень',anchor="center")
tree.heading('purity', text='Проба',anchor="center")
tree.heading('weight', text='Вес (г)',anchor="center")
tree.heading('size',text='Размер', anchor="center")
tree.heading('quantity', text='Кол-во', anchor="center")
tree.heading('price', text='Цена за шт. (₽)',anchor="center")
for col in cols:
    tree.column(col,width=100,anchor="center")

tree.grid(row=2, column=0, columnspan=2, padx=20, pady=5, sticky="nsew")
####
total_label = tk.Label(root, text="Всего товаров: 0",
                       bg=bg_color, fg=text_color,
                       font=("Segoe UI", 10, "bold"))
total_label.grid(row=3, column=0, columnspan=2, pady=(5, 0))


###базза данных
def connect_to_db():
    try:
        connection = pymysql.connect(           # +++++ заменён mysql.connector
            host="blc7gqdzfqspj5pimbbe-mysql.services.clever-cloud.com",
            port=3306,
            user="u0bp5lienenha2l1",
            password="YHsQ0s53cdT4e2lVChNF",
            database="blc7gqdzfqspj5pimbbe",
            cursorclass=pymysql.cursors.Cursor  # +++++ обычный курсор
        )
        print("DEBUG: Успешно! Подключились к Clever Cloud")
        return connection
    except Exception as e:
        print(f"Ошибка подключения к Clever Cloud: {e}")
        return None

def load_data_from_db():
    for item in tree.get_children():
        tree.delete(item)
    conn = connect_to_db()
    if not conn or not conn.open:
        return
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT CONCAT(prefix, '-', id), name, type, category, metal, gemstone, purity, weight, size, quantity, price 
                FROM jewelry_items
            """)
            rows = cursor.fetchall()
            for row in rows:
                tree.insert("", tk.END, values=row)
            total_label.config(text=f"Всего товаров: {len(rows)}")
            cached_distinct.clear()
    except Exception as e:
        print(f"Ошибка загрузки данных: {e}")
    finally:
        conn.close()
load_data_from_db()

def delete_selected_item():
    selected_item = tree.selection()

    if not selected_item:
        messagebox.showwarning("Внимание","Выберите изделия для удаления")
        return
    if not messagebox.askyesno("Удаление","Вы уверены, что хотите удалить изделие"):
        return

    item_data = tree.item(selected_item, "values")
    full_id = item_data[0]

    try:
        numeric_id = full_id.split("-")[1]
    except IndexError:
        messagebox.showerror("Ошибка","Не удалось определить ID изделия")
        return

    conn = None
    try:
        conn = connect_to_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM jewelry_items WHERE id = %s", (numeric_id,))
        conn.commit()
        cursor.close()
        messagebox.showinfo("Успех", "Изделие успешно удалено!")
        cached_distinct.clear()
        update_filters()
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось удалить из базы:\n{e}")

    finally:
        if conn and conn.open:
            conn.close()

def generate_five_digit_id():
    return random.randint(10000, 99999)

def add_item_to_db():
    data = {
        "Название": Entry_name.get(),
        "Тип": Entry_type.get(),
        "Категория": Entry_category.get(),
        "Металл": Entry_metal.get(),
        "Камень": Entry_gemstone.get(),
        "Проба": Entry_purity.get(),
        "Вес": Entry_weight.get(),
        "Размер": Entry_size.get(),
        "Количество": Entry_quantity.get(),
        "Цена": Entry_price.get()
    }

    for label, value in data.items():
        if not value.strip():
            messagebox.showwarning("Ошибка", f"Поле '{label}' обязательно!")
            return

    conn = None
    try:
        conn = connect_to_db()
        cursor = conn.cursor()
        cached_distinct.clear()
        # Определяем префикс и диапазон ID по типу
        prefix_map = {
            "Кольцо":  ("A", 1000, 1999),
            "Серьги":  ("B", 2000, 2999),
            "Браслет": ("C", 3000, 3999),
            "Цепочка": ("D", 4000, 4999)
        }
        prefix, min_id, max_id = prefix_map.get(Entry_type.get(), ("A", 1000, 1999))

        # Находим следующий свободный номер в диапазоне
        cursor.execute(
            "SELECT MAX(id) FROM jewelry_items WHERE prefix = %s AND id BETWEEN %s AND %s",
            (prefix, min_id, max_id)
        )
        row = cursor.fetchone()
        if row[0] is None:
            new_id = min_id
        else:
            if row[0] >= max_id:
                messagebox.showwarning("Лимит", f"Достигнут лимит ID для типа «{Entry_type.get()}» (макс. {max_id})")
                return
            new_id = row[0] + 1

        params = (new_id, prefix,
                  Entry_name.get(), Entry_type.get(), Entry_category.get(), Entry_metal.get(),
                  Entry_gemstone.get(), Entry_purity.get(),
                  float(Entry_weight.get().replace(',', '.') or 0),
                  Entry_size.get(), int(Entry_quantity.get() or 1),
                  float(Entry_price.get().replace(',', '.') or 0))

        cursor.execute("""
                    INSERT INTO jewelry_items (id, prefix, name, type, category, metal, gemstone, purity, weight, size, quantity, price)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, params)

        conn.commit()
        print(f"Успешно добавлено, ID = {prefix}-{new_id}")

        cursor.close()

        # Очистка полей
        Entry_name.delete(0, tk.END)
        Entry_type.set("")
        Entry_category.set("")
        Entry_metal.set("")
        Entry_gemstone.set("")
        Entry_purity.set("")
        Entry_weight.delete(0, tk.END)
        Entry_price.delete(0, tk.END)
        Entry_quantity.delete(0, tk.END)
        Entry_quantity.insert(0, "1")

        messagebox.showinfo("Успех", "Изделие добавлено!")
        load_data_from_db()

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось добавить в базу:\n{e}")


    finally:
        if conn and conn.open:
            conn.close()

# Сама кнопка
btn_add = tk.Button(border_frame, text="Добавить изделие", command=add_item_to_db, bg="#2D6A4F", fg="white")
btn_add.grid(row=3, column=0, columnspan=12, pady=20)

###выборка
def search_data(event=None):
    search = viborka.get()

    # Очищаем таблицу
    for item in tree.get_children():
        tree.delete(item)

    conn = None
    try:
        conn = connect_to_db()
        if conn and conn.open:
            cursor = conn.cursor()


            query = """
                SELECT CONCAT(prefix, '-', id), name, type, category, metal, gemstone, purity, weight, size, quantity, price 
                FROM jewelry_items 
                WHERE name LIKE %s
            """
            search_pattern = f"%{search}%"


            cursor.execute(query, (search_pattern,))

            rows = cursor.fetchall()
            for row in rows:
                tree.insert("", tk.END, values=row)

            total_label.config(text=f"Всего товаров: {len(rows)}")

            cursor.close()


    except pymysql.Error as err:
        print(f"Ошибка поиска: {err}")
        messagebox.showerror("Ошибка поиска", str(err))
    finally:
        if conn and conn.open:
            conn.close()


viborka_label=tk.Label(root,text="Поиск по названию изделия",fg=text_color,bg=bg_color,font=("Segoe UI", 12, "italic"))
viborka_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")


viborka = ttk.Entry(root,width=20)
viborka.grid(row=1, column=0, padx=220, pady=5, sticky="w")

viborka.bind("<KeyRelease>",search_data)

update = tk.Button(root, text="Нажмите для обновления данных",width=30,bg=text_color,fg="white",command=load_data_from_db)
update.grid(row=1, column=0, padx=360, pady=10, sticky="w")
current_sort_clause = ""
cached_distinct = {}

def open_buy_window():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Внимание", "Выберите изделие для покупки")
        return

    item_data = tree.item(selected_item, "values")
    item_name = item_data[1]
    price_per_item = float(item_data[10])
    available_qty = int(item_data[9])

    buy_win = tk.Toplevel(root)
    buy_win.title(f"Покупка: {item_name}")
    buy_win.geometry("350x520")
    buy_win.configure(bg=bg_color)

    tk.Label(buy_win, text=f"Изделие: {item_name}", font=("Segoe UI", 12, "bold"), bg=bg_color, fg=text_color).pack(
        pady=10)

    # Продавец
    tk.Label(buy_win, text="Продавец ФИО:", bg=bg_color, fg=text_color).pack()
    combo_seller = ttk.Combobox(buy_win, values=["Иванов И.И.", "Петров П.П.", "Сидорова А.А."], state="readonly")
    combo_seller.pack(pady=5)
    combo_seller.current(0)

    # Количество
    tk.Label(buy_win, text="Количество:", bg=bg_color, fg=text_color).pack()
    koli = tk.Entry(buy_win, width=10, justify="center",
                    validate="key", validatecommand=(v_int, '%P'))
    koli.insert(0, "1")
    koli.pack(pady=5)

    # Способ оплаты
    payment_frame = tk.LabelFrame(buy_win, text="Способ оплаты", bg=bg_color, fg=text_color, padx=10, pady=10)
    payment_frame.pack(pady=15, padx=30, fill="x")
    payment_method = tk.StringVar(value="Наличные")

    for opt in ["Наличные", "Карта", "Перевод"]:
        tk.Radiobutton(payment_frame, text=opt, variable=payment_method, value=opt, bg=bg_color,
                       selectcolor=bg_color).pack(anchor="w")

    # Метка для итоговой суммы
    total_label = tk.Label(buy_win, text=f"К оплате: {price_per_item} ₽", font=("Segoe UI", 12, "bold"),bg=bg_color, fg="#C1121F")
    total_label.pack(pady=10)

    #Функция расчета
    def calculate_total(*args):
        qty_str = koli.get()
        if qty_str.isdigit():
            total = int(qty_str) * price_per_item
            total_label.config(text=f"К оплате: {total:,.2f} ₽")
        else:
            total_label.config(text="К оплате: ошибка")

    koli.bind("<KeyRelease>", calculate_total)

    def confirm_purchase():
        qty_str = koli.get()
        if not qty_str.isdigit() or int(qty_str) <= 0:
            messagebox.showwarning("Ошибка", "Введите корректное количество!")
            return

        qty = int(qty_str)
        if qty > available_qty:
            messagebox.showwarning("Мало товара", f"В наличии только {available_qty} шт.")
            return
        current_total = qty * price_per_item
        try:
            numeric_id = item_data[0].split("-")[1]
            conn = connect_to_db()
            cursor = conn.cursor()
            cursor.execute("UPDATE jewelry_items SET quantity = quantity - %s WHERE id = %s", (qty, numeric_id))
            cursor.execute("""
                            INSERT INTO sales_history (seller_name, item_name, quantity_sold, price_per_item, payment_method, total_price) 
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (combo_seller.get(), item_name, qty, price_per_item, payment_method.get(), current_total))

            conn.commit()
            conn.close()
            messagebox.showinfo("Успех",
                                f"Покупка оформлена!\nПродавец: {combo_seller.get()}\nОплата: {payment_method.get()}")
            buy_win.destroy()
            load_data_from_db()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось обновить БД: {e}")

    tk.Button(buy_win, text="Подтвердить", command=confirm_purchase, bg=text_color, fg="white", width=20).pack(pady=10)


def set_sort(sort_clause):
    global current_sort_clause
    current_sort_clause = sort_clause
    update_filters()


def get_distinct_values(column):
    # Если есть в кэше – возвращаем сразу
    if column in cached_distinct:
        return cached_distinct[column]

    conn = connect_to_db()
    if not conn:
        return []
    try:
        cursor = conn.cursor()
        cursor.execute(
            f"SELECT DISTINCT {column} FROM jewelry_items "
            f"WHERE {column} IS NOT NULL AND {column} != '' ORDER BY {column}"
        )
        rows = cursor.fetchall()
        values = [r[0] for r in rows]
        cached_distinct[column] = values
        return values
    except Exception as e:
        print(f"Ошибка получения значений для {column}: {e}")
        return []
    finally:
        conn.close()

# Словарь для хранения BooleanVar по всем полям (кроме id)
filter_vars = {
    "name": {},
    "type": {},
    "category": {},
    "metal": {},
    "gemstone": {},
    "size": {},
    "purity": {},
    "weight": {},
    "quantity": {},
    "price": {}
}

# Функция обновления конкретного подменю
def refresh_menu(menu, category, column):
    menu.delete(0, tk.END)                      # очищаем меню
    values = get_distinct_values(column)
    current_vars = filter_vars.get(category, {})
    for val in list(current_vars.keys()):
        if val not in values:
            del current_vars[val]

    for val in values:
        if val in current_vars:
            var = current_vars[val]
        else:
            var = tk.BooleanVar()
            current_vars[val] = var
        menu.add_checkbutton(label=val, variable=var, command=update_filters)
    filter_vars[category] = current_vars

# Пост-команды для каждого подменю
def refresh_name_menu():
    refresh_menu(name_menu, "name", "name")

def refresh_type_menu():
    refresh_menu(type_menu, "type", "type")

def refresh_category_menu():
    refresh_menu(cat_menu, "category", "category")

def refresh_metal_menu():
    refresh_menu(metal_menu, "metal", "metal")

def refresh_stone_menu():
    refresh_menu(stone_menu, "gemstone", "gemstone")

def refresh_size_menu():
    refresh_menu(size_menu, "size", "size")

def refresh_purity_menu():
    refresh_menu(purity_menu, "purity", "purity")

def refresh_weight_menu():
    refresh_menu(weight_menu, "weight", "weight")

def refresh_quantity_menu():
    refresh_menu(quantity_menu, "quantity", "quantity")

def refresh_price_menu():
    refresh_menu(price_menu, "price", "price")

def update_filters():
    # 1. Собираем списки выбора из динамических переменных
    selected_names = [val for val, var in filter_vars["name"].items() if var.get()]
    selected_types = [val for val, var in filter_vars["type"].items() if var.get()]
    selected_categories = [val for val, var in filter_vars["category"].items() if var.get()]
    selected_metal = [val for val, var in filter_vars["metal"].items() if var.get()]
    selected_gemstone = [val for val, var in filter_vars["gemstone"].items() if var.get()]
    selected_sizes = [val for val, var in filter_vars["size"].items() if var.get()]
    selected_purities = [val for val, var in filter_vars["purity"].items() if var.get()]
    selected_weights = [val for val, var in filter_vars["weight"].items() if var.get()]
    selected_quantities = [val for val, var in filter_vars["quantity"].items() if var.get()]
    selected_prices = [val for val, var in filter_vars["price"].items() if var.get()]

    # 2. Формируем запрос
    query = """
    SELECT CONCAT(prefix, '-', id), name, type, category, metal, gemstone, purity, weight, size, quantity, price 
    FROM jewelry_items 
    WHERE 1=1
    """
    params = []

    if selected_names:
        placeholders = ', '.join(['%s'] * len(selected_names))
        query += f" AND name IN ({placeholders})"
        params.extend(selected_names)
    if selected_types:
        placeholders = ', '.join(['%s'] * len(selected_types))
        query += f" AND type IN ({placeholders})"
        params.extend(selected_types)
    if selected_categories:
        placeholders = ', '.join(['%s'] * len(selected_categories))
        query += f" AND category IN ({placeholders})"
        params.extend(selected_categories)
    if selected_metal:
        placeholders = ', '.join(['%s'] * len(selected_metal))
        query += f" AND metal IN ({placeholders})"
        params.extend(selected_metal)
    if selected_gemstone:
        placeholders = ', '.join(['%s'] * len(selected_gemstone))
        query += f" AND gemstone IN ({placeholders})"
        params.extend(selected_gemstone)
    if selected_sizes:
        placeholders = ', '.join(['%s'] * len(selected_sizes))
        query += f" AND size IN ({placeholders})"
        params.extend(selected_sizes)
    if selected_purities:
        placeholders = ', '.join(['%s'] * len(selected_purities))
        query += f" AND purity IN ({placeholders})"
        params.extend(selected_purities)
    if selected_weights:
        placeholders = ', '.join(['%s'] * len(selected_weights))
        query += f" AND weight IN ({placeholders})"
        params.extend(selected_weights)
    if selected_quantities:
        placeholders = ', '.join(['%s'] * len(selected_quantities))
        query += f" AND quantity IN ({placeholders})"
        params.extend(selected_quantities)
    if selected_prices:
        placeholders = ', '.join(['%s'] * len(selected_prices))
        query += f" AND price IN ({placeholders})"
        params.extend(selected_prices)

    # Сортировка
    if current_sort_clause:
        query += f" ORDER BY {current_sort_clause}"

    # 3. Выполняем
    conn = None
    try:
        conn = connect_to_db()
        if conn and conn.open:
            cursor = conn.cursor()
            cursor.execute(query, tuple(params))
            rows = cursor.fetchall()

            for item in tree.get_children():
                tree.delete(item)
            for row in rows:
                tree.insert("", tk.END, values=row)

            total_label.config(text=f"Всего товаров: {len(rows)}")

            cursor.close()
    except pymysql.Error as err:
            print(f"Ошибка фильтрации: {err}")
    finally:
        if conn and conn.open:
            conn.close()

def reset_filters():
    viborka.delete(0, tk.END)
    for cat_dict in filter_vars.values():
        for var in cat_dict.values():
            var.set(False)
    update_filters()


def reset_sorting():
    global current_sort_clause
    current_sort_clause = ""
    update_filters()

# ====== МЕНЮ ======
Menu = tk.Menu(root)
root.config(menu=Menu)

filter_menu = tk.Menu(Menu, tearoff=0, fg=text_color)
Menu.add_cascade(label="Фильтры", menu=filter_menu)

# Подменю с динамическим наполнением
name_menu = tk.Menu(filter_menu, tearoff=0, postcommand=refresh_name_menu)
filter_menu.add_cascade(label="По названию", menu=name_menu)

type_menu = tk.Menu(filter_menu, tearoff=0, postcommand=refresh_type_menu)
filter_menu.add_cascade(label="По типу", menu=type_menu)

cat_menu = tk.Menu(filter_menu, tearoff=0, postcommand=refresh_category_menu)
filter_menu.add_cascade(label="По категории", menu=cat_menu)

metal_menu = tk.Menu(filter_menu, tearoff=0, postcommand=refresh_metal_menu)
filter_menu.add_cascade(label="По металлу", menu=metal_menu)

stone_menu = tk.Menu(filter_menu, tearoff=0, postcommand=refresh_stone_menu)
filter_menu.add_cascade(label="По камню", menu=stone_menu)

size_menu = tk.Menu(filter_menu, tearoff=0, postcommand=refresh_size_menu)
filter_menu.add_cascade(label="По размеру", menu=size_menu)

purity_menu = tk.Menu(filter_menu, tearoff=0, postcommand=refresh_purity_menu)
filter_menu.add_cascade(label="По пробе", menu=purity_menu)

weight_menu = tk.Menu(filter_menu, tearoff=0, postcommand=refresh_weight_menu)
filter_menu.add_cascade(label="По весу", menu=weight_menu)

quantity_menu = tk.Menu(filter_menu, tearoff=0, postcommand=refresh_quantity_menu)
filter_menu.add_cascade(label="По количеству", menu=quantity_menu)

price_menu = tk.Menu(filter_menu, tearoff=0, postcommand=refresh_price_menu)
filter_menu.add_cascade(label="По цене", menu=price_menu)

# Кнопка сброса фильтров
filter_menu.add_separator()
filter_menu.add_command(label="Сбросить фильтры", command=reset_filters, foreground="red")

#сортировки
sort_menu = tk.Menu(Menu, tearoff=0)
Menu.add_cascade(label="Сортировки", menu=sort_menu)
sort_menu.add_command(label="Сортировка от А до Я",command=lambda: set_sort("name ASC"))
sort_menu.add_command(label="Сортировка от Я до А",command=lambda: set_sort("name DESC"))
sort_menu.add_command(label="Сначала дешевые",command=lambda: set_sort("price ASC"))

sort_menu.add_separator()
sort_menu.add_command(label="Сбросить сортировку", command=reset_sorting,foreground="red")

delete = tk.Button(root, text="Удалить",bg="#2D6A4F", fg="white",width=25,command=delete_selected_item)
delete.grid(row=1, column=0, padx=590, pady=10, sticky="w")

def show_manual():
    messagebox.showinfo("Руководство пользователя",
        "1. Добавление: Заполните поля и нажмите 'Добавить изделие'.\n"
        "2. Поиск: Вводите название в строку поиска для мгновенной фильтрации.\n"
        "3. Покупка: Выберите товар в таблице и нажмите 'Купить изделие'.\n"
        "4. Отчеты: Нажмите кнопку 'Отчеты' для просмотра статистики продаж.")

def show_about_program():
    messagebox.showinfo("О программе",
        "Система учета продаж ювелирных изделий 'Алмазный путь'\n\n"
        "Программный комплекс предназначен для автоматизации розничной торговли: "
        "учета товарных запасов, фиксации продаж и ведения аналитики покупательского спроса.\n\n"
        "Версия: 3.0\n"
        "Технологии: Python (Tkinter), MySQL (СУБД).")

def show_me():
    messagebox.showinfo("О разработчике","Исаев Рамин Захид оглы ИС-943")

inf_menu = tk.Menu(Menu, tearoff=0)
Menu.add_cascade(label="Справочная информация", menu=inf_menu)
inf_menu.add_command(label="Руководство пользователя",command=show_manual)
inf_menu.add_command(label="О программе",command=show_about_program)
inf_menu.add_command(label="О разработчике",command=show_me)


#покупка
buy = tk.Button(root,text="Купить изделие",bg="#2D6A4F", fg="white",command=open_buy_window)
buy.grid(row=1, column=0, padx=779, pady=10, sticky="w")

#отчеты
def new_window():
    root2 = tk.Toplevel(root)
    root2.title("Отчеты")
    root2.state("zoomed")
    root2.configure(bg=bg_color)
    root2.columnconfigure(0, weight=1)
    root2.rowconfigure(1, weight=1)

    # Панель фильтров
    frame = tk.Frame(root2, bg=bg_color)
    frame.grid(row=0, column=0, sticky="ew", padx=20, pady=10)


    def create_filter_combo(parent, label_text, values):
        tk.Label(parent, text=label_text, bg=bg_color).pack(side="left", padx=(10, 5))
        combo = ttk.Combobox(parent, values=values, state="readonly", width=15)
        combo.pack(side="left", padx=5)
        return combo

    # Дата и Продавец
    tk.Label(frame, text="С:", bg=bg_color).pack(side="left", padx=5)
    date_from = DateEntry(frame, width=10, date_pattern='yyyy-mm-dd')
    date_from.pack(side="left", padx=5)

    tk.Label(frame, text="По:", bg=bg_color).pack(side="left", padx=5)
    date_to = DateEntry(frame, width=10, date_pattern='yyyy-mm-dd')
    date_to.pack(side="left", padx=5)

    def reset_all_filters():
        combo_seller.set("— Все —")
        po_kat.set("— Все —")
        po_mat.set("— Все —")

        date_from.set_date(date.today())
        date_to.set_date(date.today())

        load_report_data()

    tk.Label(frame, text="Продавец:", bg=bg_color).pack(side="left", padx=5)
    combo_seller = ttk.Combobox(frame, values=["Иванов И.И.", "Петров П.П.", "Сидорова А.А."], state="readonly",width=15)
    combo_seller.set("— Все —")
    combo_seller.pack(side="left", padx=5)

    # Категории и Материалы
    po_kat = create_filter_combo(frame, "Категория:", ["Свадебные", "Женские", "Мужские", "Детские"])
    po_kat.set("— Все —")
    po_mat = create_filter_combo(frame, "Материал:",
                                 ["Золото", "Красное золото", "Белое золото", "Желтое золото", "Серебро"])
    po_mat.set("— Все —")
    #популярность
    def load_popular_data():
        start_date = date_from.get()
        end_date = date_to.get()
        for i in tree_report.get_children(): tree_report.delete(i)

        try:
            conn = connect_to_db()
            cursor = conn.cursor(pymysql.cursors.DictCursor)

            # Группируем по уникальному ID изделия, чтобы база не выдавала ошибку 1055
            query = """
                            SELECT 
                                CONCAT(MAX(j.prefix), '-', MIN(j.id)) as full_id, 
                                DATE_FORMAT(MAX(s.sale_date), '%%Y-%%m-%%d') as sale_date,  -- Вот здесь магия
                                '—' as seller_name, 
                                s.item_name, 
                                j.metal, 
                                j.category, 
                                SUM(s.quantity_sold) as total_qty, 
                                AVG(s.price_per_item) as price, 
                                '—' as payment_method, 
                                SUM(s.total_price) as total_price
                            FROM sales_history s
                            JOIN jewelry_items j ON s.item_name = j.name
                            WHERE DATE(s.sale_date) BETWEEN %s AND %s
                            GROUP BY j.id, s.item_name, j.metal, j.category
                            ORDER BY total_qty DESC
                        """
            cursor.execute(query, (start_date, end_date))
            rows = cursor.fetchall()

            total_sum = 0
            total_qty = 0

            for row in rows:
                tree_report.insert("", tk.END, values=(
                    row['full_id'],
                    row['sale_date'],
                    row['seller_name'],
                    row['item_name'],
                    row['metal'] or "—",
                    row['category'] or "—",
                    row['total_qty'],
                    f"{row['price']:.2f}",
                    row['payment_method'],
                    f"{row['total_price']:.2f}"
                ))
                total_sum += row['total_price']
                total_qty += row['total_qty']
            lbl_summ.config(text=f"💰 Выручка: {total_sum:,.2f} руб. | 📦 Всего: {total_qty} шт.")
            conn.close()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка БД: {e}")

    # Таблица
    cols = ('ID', 'Дата', 'Продавец', 'Товар','Материал', 'Категория','Кол-во проданных', 'Цена', 'Оплата',"Итого")
    tree_report = ttk.Treeview(root2, columns=cols, show='headings')
    for col in cols:
        tree_report.heading(col, text=col)
        tree_report.column(col, width=80, anchor='center')
    tree_report.grid(row=1, column=0, sticky="nsew", padx=20, pady=5)

    # Инфо-панель
    summ_frame = tk.Frame(root2, bg="#2D6A4F")
    summ_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=10)
    lbl_summ = tk.Label(summ_frame, text="Загрузка данных...", font=("Segoe UI", 12, "bold"), bg="#2D6A4F", fg="white")
    lbl_summ.pack(pady=10)

    def load_efficiency():
        for i in tree_report.get_children(): tree_report.delete(i)

        start_date = date_from.get_date().strftime('%Y-%m-%d')
        end_date = date_to.get_date().strftime('%Y-%m-%d')

        try:
            conn = connect_to_db()
            cursor = conn.cursor(pymysql.cursors.DictCursor)

            # Запрос группирует данные именно по продавцам
            query = """
                SELECT 
                    seller_name, 
                    COUNT(*) as sales_count, 
                    SUM(quantity_sold) as total_qty, 
                    SUM(total_price) as total_revenue
                FROM sales_history
                WHERE DATE(sale_date) BETWEEN %s AND %s
                GROUP BY seller_name
                ORDER BY total_revenue DESC
            """
            cursor.execute(query, (start_date, end_date))
            rows = cursor.fetchall()

            for row in rows:
                tree_report.insert("", tk.END, values=(
                    "—", "—", row['seller_name'],
                    "—", "—", "—", row['total_qty'],
                    "—", "—", f"{row['total_revenue']:.2f}"
                ))

            lbl_summ.config(text="📊 Отчет по эффективности сотрудников")
            conn.close()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка БД: {e}")



    # Логика загрузки
    def load_report_data(is_popular=False):
        for i in tree_report.get_children(): tree_report.delete(i)

        # Берем значения напрямую из объектов DateEntry
        start_date = date_from.get_date().strftime('%Y-%m-%d')
        end_date = date_to.get_date().strftime('%Y-%m-%d')

        seller = combo_seller.get()
        category = po_kat.get()
        material = po_mat.get()

        try:
            conn = connect_to_db()
            cursor = conn.cursor(pymysql.cursors.DictCursor)

            if is_popular:
                query = """
                    SELECT 
                        CONCAT(MAX(j.prefix), '-', MIN(j.id)) as full_id, 
                        '—' as sale_date, 
                        '—' as seller_name, 
                        s.item_name, 
                        j.metal, 
                        j.category, 
                        SUM(s.quantity_sold) as quantity_sold, 
                        AVG(s.price_per_item) as price_per_item, 
                        '—' as payment_method, 
                        SUM(s.total_price) as total_price
                    FROM sales_history s
                    JOIN jewelry_items j ON s.item_name = j.name
                    WHERE DATE(s.sale_date) BETWEEN %s AND %s
                """
            else:
                query = """
                    SELECT 
                        CONCAT(j.prefix, '-', j.id) as full_id, 
                        DATE_FORMAT(s.sale_date, '%%Y-%%m-%%d') as sale_date,
                        s.seller_name, s.item_name, j.metal, j.category, 
                        s.quantity_sold, s.price_per_item, s.payment_method, s.total_price
                    FROM sales_history s
                    LEFT JOIN jewelry_items j ON s.item_name = j.name
                    WHERE DATE(s.sale_date) BETWEEN %s AND %s
                """

            params = [start_date, end_date]

            if seller != "— Все —":
                query += " AND s.seller_name = %s"
                params.append(seller)
            if category != "— Все —":
                query += " AND j.category = %s"
                params.append(category)
            if material != "— Все —":
                query += " AND j.metal = %s"
                params.append(material)

            if is_popular:
                query += " GROUP BY j.id, s.item_name, j.metal, j.category ORDER BY quantity_sold DESC"

            cursor.execute(query, tuple(params))
            rows = cursor.fetchall()

            total_sum = 0
            total_qty = 0

            for row in rows:

                cena_chislo = float(row['price_per_item'] or 0)
                cena_krasivaya = f"{cena_chislo:.2f}"


                itogo_chislo = float(row['total_price'] or 0)
                itogo_krasivo = f"{itogo_chislo:.2f}"


                tree_report.insert("", tk.END, values=(
                    row['full_id'], row['sale_date'], row['seller_name'],
                    row['item_name'], row['metal'] or "—", row['category'] or "—",
                    row['quantity_sold'], cena_krasivaya,
                    row['payment_method'], itogo_krasivo
                ))
                total_sum += itogo_chislo
                total_qty += int(row['quantity_sold'] or 0)

            lbl_summ.config(text=f"💰 Выручка: {total_sum:,.2f} руб. | 📦 Всего: {total_qty} шт.")
            conn.close()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка БД: {e}")

    # При выборе в комбобоксах — сразу обновляем отчет
    po_kat.bind("<<ComboboxSelected>>", lambda e: load_report_data())
    po_mat.bind("<<ComboboxSelected>>", lambda e: load_report_data())
    combo_seller.bind("<<ComboboxSelected>>", lambda e: load_report_data())

    # Привязка для календаря (обновляется при смене даты)
    date_from.bind("<<DateEntrySelected>>", lambda e: load_report_data())
    date_to.bind("<<DateEntrySelected>>", lambda e: load_report_data())
    # сохранение в ворд
    # сохранение в ворд
    def save():
        all_data = []
        for row_id in tree_report.get_children():
            all_data.append(tree_report.item(row_id)["values"])

        if not all_data:
            messagebox.showwarning("Внимание", "Таблица пуста!")
            return

        # Определяем колонки, которые полностью состоят из "—"
        exclude_indices = []
        for col_idx in range(len(cols)):
            if all(row[col_idx] == "—" for row in all_data):
                exclude_indices.append(col_idx)

        # Создаём Excel-книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт"

        # Заголовки (без исключённых колонок)
        headers = [h for i, h in enumerate(cols) if i not in exclude_indices]
        ws.append(headers)

        # Данные (без исключённых колонок)
        for row in all_data:
            filtered_row = [val for i, val in enumerate(row) if i not in exclude_indices]
            ws.append(filtered_row)

        # Автоматическая ширина колонок
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

        # Сохраняем файл
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            try:
                wb.save(file_path)
                messagebox.showinfo("Успех", "Отчёт успешно сохранён в Excel!")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")

    # Кнопки
    po_pop = tk.Button(frame, text="По популярности", bg="#2D6A4F", fg="white",
                       command=lambda: load_report_data(is_popular=True))
    po_pop.pack(side="left", padx=5)

    btn_eff = tk.Button(frame, text="Эффективность", bg="#2D6A4F", fg="white",
                        command=load_efficiency)
    btn_eff.pack(side="left", padx=5)

    btn_apply = tk.Button(frame, text="Обновить", bg=text_color, fg="white",
                          command=lambda: load_report_data(is_popular=False))
    btn_apply.pack(side="left", padx=5)

    btn_reset = tk.Button(frame, text="Сбросить всё", bg="#8B0000", fg="white",
                          command=reset_all_filters)
    btn_reset.pack(side="left", padx=5)

    btn_save = tk.Button(frame, text="Сохранить в Excel", bg="#2D6A4F", fg="white", command=save)
    btn_save.pack(side="left", padx=5)



    # 3. Итоги
    summ_frame = tk.Frame(root2, bg="#2D6A4F")
    summ_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=10)

    lbl_summ = tk.Label(summ_frame, text="", font=("Segoe UI", 12, "bold"), bg="#2D6A4F", fg="white")
    lbl_summ.pack(side="left", padx=10)

    load_report_data()



btn2 = tk.Button(root, text ="Отчеты",bg="#2D6A4F", fg="white",width=20,command=new_window)
btn2.grid(row=1, column=0, padx=885, pady=10, sticky="w")

root.mainloop()